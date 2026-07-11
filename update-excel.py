import openpyxl
import json
import os
import sys

def main():
    print("🐍 Python Excel Writer executing...")
    excel_path = "Database.xlsx"
    updates_path = "temp-updates.json"
    
    if not os.path.exists(excel_path):
        print(f"❌ Error: {excel_path} not found")
        sys.exit(1)
        
    if not os.path.exists(updates_path):
        print(f"❌ Error: {updates_path} not found")
        sys.exit(1)
        
    with open(updates_path, "r", encoding="utf-8") as f:
        updates = json.load(f)
        
    if not updates:
        print("ℹ️ No updates to write.")
        sys.exit(0)
        
    print(f"📦 Loaded {len(updates)} row updates to apply.")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb["sample2"]
        
        # Identify headers in row 1
        headers = [cell.value for cell in sheet[1]]
        print("Headers found in sheet:", headers)
        
        # Dynamically locate columns to match headers precisely
        try:
            topic_col = headers.index("Topic") + 1
        except ValueError:
            print("❌ Error: 'Topic' column not found.")
            sys.exit(1)
            
        try:
            status_col = headers.index("Status") + 1
        except ValueError:
            # If Status doesn't exist, append it
            status_col = len(headers) + 1
            sheet.cell(row=1, column=status_col, value="Status")
            headers.append("Status")
            print(f"➕ Created missing column 'Status' in column {status_col}")
            
        try:
            talkby_col = headers.index("Talk By") + 1
        except ValueError:
            # If Talk By doesn't exist, append it
            talkby_col = len(headers) + 1
            sheet.cell(row=1, column=talkby_col, value="Talk By")
            headers.append("Talk By")
            print(f"➕ Created missing column 'Talk By' in column {talkby_col}")
            
        # Apply row updates
        for item in updates:
            row_num = item["rowNumber"]
            print(f"✍️ Updating Row {row_num} (Status: {item['status']}, Talk By: {item['talkBy']})")
            
            # Write cells safely
            sheet.cell(row=row_num, column=topic_col, value=item["topic"])
            sheet.cell(row=row_num, column=status_col, value=item["status"])
            sheet.cell(row=row_num, column=talkby_col, value=item["talkBy"])
            
        wb.save(excel_path)
        print("✅ Database.xlsx successfully updated and saved by Python openpyxl!")
        
        # Synchronize CSV files as well
        import csv
        def sync_to_csv(dest_csv_path):
            with open(dest_csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in sheet.iter_rows(values_only=True):
                    if any(row):
                        writer.writerow(row[:6])
            print(f"✅ CSV synchronized to {dest_csv_path}!")

        sync_to_csv("Database.csv")
        
        # Also copy/save to public/Database.xlsx and public/Database.csv for client-side serving in Vite
        public_dir = "public"
        if os.path.exists(public_dir):
            public_excel_path = os.path.join(public_dir, "Database.xlsx")
            wb.save(public_excel_path)
            print("✅ public/Database.xlsx successfully synchronized!")
            sync_to_csv(os.path.join(public_dir, "Database.csv"))
            
        # Crucial: Also copy/save to dist/Database.xlsx and dist/Database.csv if dist exists (production container serves from dist)
        dist_dir = "dist"
        if os.path.exists(dist_dir):
            dist_excel_path = os.path.join(dist_dir, "Database.xlsx")
            wb.save(dist_excel_path)
            print("✅ dist/Database.xlsx successfully synchronized!")
            sync_to_csv(os.path.join(dist_dir, "Database.csv"))
        
    except Exception as e:
        print(f"❌ Fatal Python error updating Excel: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
